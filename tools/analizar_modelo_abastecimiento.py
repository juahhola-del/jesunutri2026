import csv
import json
import math
import os
import re
import statistics
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "abastecimiento_modelo"
OUT_DIR.mkdir(parents=True, exist_ok=True)

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "")

MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
}
PAC_MONTH_COLUMNS = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}
MANUAL_ALIASES = {
    "althera": "althera",
    "jalea con azucar": "jalea normal",
    "jalea diet 500 gr": "jalea dietetica",
    "chuno 500 gr": "chuno",
    "fideos": "fideos tallarin",
    "tallarines": "fideos tallarin",
    "espirales": "fideos guiso espirales",
    "alusa": "papel film",
    "vasos": "vaso termico",
    "similac total confort": "total confort",
    "formula infantil semi elemental total confort": "total confort",
}


def normalize(value):
    text = str(value or "").lower()
    text = (
        text.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def number(value):
    if value is None or hasattr(value, "year"):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:[\.,]\d+)?", str(value))
    return float(match.group(0).replace(",", ".")) if match else 0.0


def month_from_filename(path):
    name = path.name.upper()
    for label, month in MONTHS.items():
        if label in name:
            return month
    return None


def supabase_get(table, select, params=None):
    if not SUPABASE_URL or not SUPABASE_ANON_KEY:
        raise RuntimeError("Configura SUPABASE_URL y SUPABASE_ANON_KEY para consultar Supabase.")
    query = {"select": select}
    if params:
        query.update(params)
    url = f"{SUPABASE_URL}/rest/v1/{table}?{urllib.parse.urlencode(query, doseq=True)}"
    request = urllib.request.Request(
        url,
        headers={
            "apikey": SUPABASE_ANON_KEY,
            "Authorization": f"Bearer {SUPABASE_ANON_KEY}",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def parse_pac():
    path = ROOT / "PAC ALIMENTACION 2026_REVISADO.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["TRABAJAR"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = row[3]
        detail = row[4]
        if not code or not detail:
            continue
        months = {m: number(row[6 + m]) for m in range(1, 13)}
        rows.append(
            {
                "code": str(code).strip(),
                "name": str(detail).strip(),
                "norm": normalize(detail),
                "annual": number(row[6]),
                "unit_price": number(row[5]),
                "months": months,
            }
        )
    return rows


def parse_monthly_orders():
    rows = []
    for path in sorted((ROOT / "pedidos").glob("PEDIDO * 2026 CENTRAL.xlsx")):
        month = month_from_filename(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                product = code = unit = None
                qty = 0.0
                if sheet_name == "Pedido Abarrotes" and len(row) > 5 and row[1] and row[2] and normalize(row[1]) != "codigo":
                    code, product, unit, qty = row[1], row[2], row[3], number(row[5])
                elif sheet_name == "Pedido Desechables" and len(row) > 6 and row[2] and row[3] and normalize(row[2]) != "codigo":
                    code, product, unit, qty = row[2], row[3], "unidad", number(row[6])
                elif sheet_name in ("ENTERALES", "INFANTIL") and len(row) > 6 and row[2] and row[3] and "codigo" not in normalize(row[2]):
                    code, product, unit, qty = row[2], row[3], row[4], number(row[6])
                elif sheet_name == "Pedido verduras" and len(row) > 5 and row[1] and number(row[5]) > 0 and not re.search("seccion|pedido|geonox|mensual", normalize(row[1])):
                    product, unit, qty = row[1], "unidad", number(row[5])
                elif sheet_name == "PEDIDO DE VERDURAS" and len(row) > 7 and row[3] and number(row[7]) > 0 and "pedido" not in normalize(row[3]):
                    product, unit, qty = row[3], "unidad", number(row[7])
                elif sheet_name == "Calendario Recepciones":
                    vals = list(row[:10])
                    text = " ".join(normalize(x) for x in vals if x is not None)
                    if "huevos" in text and any(str(x or "").startswith("ALIM-0029") for x in vals):
                        idx = [normalize(x) for x in vals].index("huevos")
                        nums = [number(x) for x in vals[:idx] if number(x) > 0]
                        code, product, unit, qty = "ALIM-0029", "Huevos", "unidad", (nums[-1] if nums else 0)
                    elif "carne molida de pavo" in text and any(str(x or "").startswith("ALIM-0139") for x in vals):
                        idx = [normalize(x) for x in vals].index("carne molida de pavo")
                        nums = [number(x) for x in vals[:idx] if number(x) > 0]
                        code, product, unit, qty = "ALIM-0139", "Carne molida de pavo", "unidad", (nums[-1] if nums else 0)
                if product and qty > 0:
                    rows.append(
                        {
                            "month": month,
                            "sheet": sheet_name,
                            "code": str(code or "").replace(" ", "").upper(),
                            "name": str(product).strip(),
                            "norm": normalize(product),
                            "unit": str(unit or "unidad").strip(),
                            "quantity": qty,
                            "source_file": path.name,
                        }
                    )
    return rows


def parse_daily_may_enterals():
    rows = []
    for path in sorted((ROOT / "Planillas Mayo").glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in ("ENTERALES", "MODULARES DIARIAS"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if sheet_name == "ENTERALES":
                for row in ws.iter_rows(min_row=4, values_only=True):
                    product = row[1] if len(row) > 1 else None
                    if not product:
                        continue
                    qty = sum(number(cell) for cell in row[3:24])
                    if qty > 0:
                        rows.append({"name": str(product).strip(), "norm": normalize(product), "quantity": qty, "source_file": path.name, "sheet": sheet_name})
            else:
                for row in ws.iter_rows(min_row=4, values_only=True):
                    for base in (3, 12):
                        product = row[base] if len(row) > base else None
                        qty = number(row[base + 2]) if len(row) > base + 2 else 0
                        if product and qty > 0:
                            rows.append({"name": str(product).strip(), "norm": normalize(product), "quantity": qty, "source_file": path.name, "sheet": sheet_name})
    return rows


def best_match(product, candidates):
    pname = product["nombre_normalizado"] or normalize(product["nombre"])
    alias = MANUAL_ALIASES.get(pname)
    best = None
    best_score = 0
    product_words = set(pname.split())
    for candidate in candidates:
        cname = candidate["norm"]
        if not cname:
            continue
        if alias and alias in cname:
            score = 98
        elif pname == cname:
            score = 100
        elif (pname in cname or cname in pname) and min(len(pname.split()), len(cname.split())) >= 2:
            score = 86
        else:
            words = set(cname.split())
            shared = len(product_words & words)
            score = shared / max(1, len(product_words | words)) * 70
        if score > best_score:
            best = candidate
            best_score = score
    return best if best_score >= 42 else None, best_score


def p75(values):
    if not values:
        return 0.0
    ordered = sorted(values)
    return ordered[min(len(ordered) - 1, math.ceil(len(ordered) * 0.75) - 1)]


def main():
    products = supabase_get(
        "productos_insumos",
        "id,nombre,nombre_normalizado,unidad_default,stock_minimo,consumo_promedio_diario,critico,activo",
        {"activo": "eq.true"},
    )
    inventory = supabase_get(
        "inventario_lotes_disponibles",
        "producto_id,nombre,cantidad_disponible,unidad,fecha_vencimiento,lote,activo,deleted_at",
        {"activo": "eq.true", "cantidad_disponible": "gt.0"},
    )
    stock_by_product = defaultdict(float)
    for row in inventory:
        stock_by_product[row["producto_id"]] += float(row.get("cantidad_disponible") or 0)

    pac_rows = parse_pac()
    order_rows = parse_monthly_orders()
    daily_rows = parse_daily_may_enterals()

    pac_by_code = {row["code"].replace(" ", "").upper(): row for row in pac_rows}
    orders_by_code = defaultdict(list)
    for row in order_rows:
        if row["code"]:
            orders_by_code[row["code"]].append(row)

    result_rows = []
    for product in products:
        pname = product["nombre"]
        product_norm = product["nombre_normalizado"] or normalize(pname)
        pac_match, pac_score = best_match(product, pac_rows)
        order_candidates = order_rows
        if pac_match and pac_match["code"].replace(" ", "").upper() in orders_by_code:
            order_candidates = orders_by_code[pac_match["code"].replace(" ", "").upper()]
        order_match, order_score = best_match(product, order_candidates)

        monthly_values = []
        matched_order_names = set()
        if order_match:
            if order_match.get("code") and order_match["code"] in orders_by_code:
                relevant = orders_by_code[order_match["code"]]
            else:
                relevant = [row for row in order_rows if row["norm"] == order_match["norm"] or product_norm in row["norm"] or row["norm"] in product_norm]
            monthly_by_month = defaultdict(float)
            for row in relevant:
                if row["month"]:
                    monthly_by_month[row["month"]] += row["quantity"]
                    matched_order_names.add(row["name"])
            monthly_values = [monthly_by_month[m] for m in sorted(monthly_by_month) if monthly_by_month[m] > 0]

        pac_next = pac_match["months"].get(7, 0) if pac_match else 0
        pac_monthly_values = [v for v in (pac_match["months"].values() if pac_match else []) if v > 0]
        daily_matches = [row for row in daily_rows if product_norm in row["norm"] or row["norm"] in product_norm]
        daily_may_total = sum(row["quantity"] for row in daily_matches)
        daily_projected_month = daily_may_total if daily_may_total > 0 else 0

        monthly_avg = statistics.mean(monthly_values) if monthly_values else 0.0
        monthly_p75 = p75(monthly_values)
        pac_avg = statistics.mean(pac_monthly_values) if pac_monthly_values else 0.0

        if monthly_values:
            base_monthly = monthly_avg
            source = "pedidos_mensuales"
            confidence = "alta" if len(monthly_values) >= 4 else "media"
        elif pac_avg:
            base_monthly = pac_avg
            source = "pac_2026"
            confidence = "media-baja"
        else:
            base_monthly = 0.0
            source = "sin_data"
            confidence = "baja"

        safety_min = math.ceil(max(base_monthly * 0.25, monthly_p75 * 0.25, base_monthly / 30.4375 * 7)) if base_monthly > 0 else 0
        current_min = float(product.get("stock_minimo") or 0)
        final_min = safety_min if base_monthly > 0 else current_min
        daily_consumption = round(base_monthly / 30.4375, 3) if base_monthly > 0 else 0
        stock_actual = stock_by_product[product["id"]]
        next_month_consumption = pac_next if source == "pac_2026" and pac_next > 0 else base_monthly
        suggested_order = max(0, math.ceil(next_month_consumption + final_min - stock_actual))

        result_rows.append(
            {
                "producto_id": product["id"],
                "producto": pname,
                "stock_actual": round(stock_actual, 3),
                "minimo_actual": current_min,
                "minimo_sugerido": final_min,
                "consumo_diario_sugerido": daily_consumption,
                "pedido_julio_sugerido": suggested_order,
                "base_mensual": round(base_monthly, 3),
                "promedio_pedidos": round(monthly_avg, 3),
                "p75_pedidos": round(monthly_p75, 3),
                "pac_julio": round(pac_next, 3),
                "demanda_diaria_mayo": round(daily_projected_month, 3),
                "fuente": source,
                "confianza": confidence,
                "match_pac": pac_match["name"] if pac_match else "",
                "match_pedido": " | ".join(sorted(matched_order_names))[:240],
                "regla": "min=max(min_actual, 25% promedio pedido, 25% p75 pedido, 7 dias consumo); pedido=max(0, base_mensual + minimo - stock_actual)",
            }
        )

    result_rows.sort(key=lambda row: (row["fuente"] == "sin_data", row["producto"].lower()))

    csv_path = OUT_DIR / "minimos_y_prediccion_julio_2026.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(result_rows[0].keys()))
        writer.writeheader()
        writer.writerows(result_rows)

    sql_rows = [row for row in result_rows if row["minimo_sugerido"] > 0 or row["consumo_diario_sugerido"] > 0]
    sql_path = ROOT / "supabase" / "25_aplicar_minimos_y_prediccion_julio_2026.sql"
    with sql_path.open("w", encoding="utf-8") as fh:
        fh.write("-- JESUnutri - aplicar minimos sugeridos y revisar pedido mensual julio 2026.\n")
        fh.write("-- Fuente: PAC Alimentacion 2026, pedidos mensuales ene-jun, demanda diaria mayo trazable y stock real Supabase.\n")
        fh.write("begin;\n\n")
        fh.write("with sugerencias(producto_id, stock_minimo, consumo_promedio_diario, pedido_julio_sugerido, fuente, confianza) as (values\n")
        value_lines = []
        for row in sql_rows:
            value_lines.append(
                "  ('{producto_id}'::uuid, {minimo_sugerido}, {consumo_diario_sugerido}, {pedido_julio_sugerido}, '{fuente}', '{confianza}')".format(
                    **row
                )
            )
        fh.write(",\n".join(value_lines))
        fh.write("\n),\n")
        fh.write(
            "actualizados as (\n"
            "  update productos_insumos p\n"
            "  set stock_minimo = s.stock_minimo,\n"
            "      consumo_promedio_diario = s.consumo_promedio_diario,\n"
            "      updated_at = now()\n"
            "  from sugerencias s\n"
            "  where p.id = s.producto_id\n"
            "    and (s.stock_minimo > 0 or s.consumo_promedio_diario > 0)\n"
            "  returning p.id\n"
            ")\n"
            "select p.nombre, s.stock_minimo, s.consumo_promedio_diario, s.pedido_julio_sugerido, s.fuente, s.confianza\n"
            "from sugerencias s\n"
            "join productos_insumos p on p.id = s.producto_id\n"
            "join actualizados a on a.id = p.id\n"
            "order by p.nombre;\n\n"
            "commit;\n"
        )

    summary = {
        "products": len(result_rows),
        "with_monthly_orders": sum(1 for row in result_rows if row["fuente"] == "pedidos_mensuales"),
        "with_daily": sum(1 for row in result_rows if row["fuente"] == "demanda_diaria_mayo"),
        "with_pac_only": sum(1 for row in result_rows if row["fuente"] == "pac_2026"),
        "without_data": sum(1 for row in result_rows if row["fuente"] == "sin_data"),
        "csv": str(csv_path),
        "sql": str(sql_path),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
